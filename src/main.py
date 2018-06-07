import logging
import openpyxl
from openpyxl.styles import colors
from openpyxl import Workbook
from openpyxl.styles import Font, Color, Alignment, PatternFill


class VarianzaFormat:

    class Options:
        adjust_content = True
        font = None
        header = 1
        header_bg = None
        odd_bg = None
        even_bg = None
        thousand_separator = True
        input_filepath = './res/input.xlsx'
        output_filepath = './res/input_generated.xlsx'

    def __init__(self):
        self.options = self.Options()

    def run(self):
        # print("Formatting workbook")
        logging.info('Applying format ...')
        logging.debug('adjust_content %r', self.options.adjust_content)
        logging.debug('font :      %s', self.options.font)
        logging.debug('header :    %d', self.options.header)
        logging.debug('header_bg : %s', self.options.header_bg)
        logging.debug('odd_bg :    %s', self.options.odd_bg)
        logging.debug('even_bg :   %s', self.options.even_bg)
        logging.debug('thousand_separator : %r', self.options.thousand_separator)
        wb = openpyxl.load_workbook(filename=self.options.input_filepath)

        # Select active Worksheet
        ws = wb.active

        # 1. Change cell font
        if self.options.font is not None:
            for row in ws.iter_rows():
                for cell in row:
                    cell.font = Font(name=self.options.font)

        # 1. Adjust content
        # See https://stackoverflow.com/questions/13197574/openpyxl-adjust-column-width-size/14450572
        if self.options.adjust_content:
            for col in ws.columns:
                max_length = 0
                column = col[0].column  # Get the column name

                for cell in col:
                    try:  # Necessary to avoid error on empty cells
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except Exception:
                        pass
                adjusted_width = (max_length + 2) * 2
                ws.column_dimensions[column].width = adjusted_width

        # 3. Update cell background (header, odd,even)
        for row in ws.iter_rows():

            bg = 'FFFFFF'
            row_number = row[0].row
            if self.options.header == row_number:
                bg = self.options.header_bg
            else:
                if row_number % 2 == 0:
                    bg = self.options.even_bg  # Even
                else:
                    bg = self.options.odd_bg  # Odd
            for cell in row:
                if bg is not None:
                    cell.fill = PatternFill(fgColor=bg, fill_type="solid")

        # 4. Add thousands separator
        if self.options.thousand_separator:
            for row in ws.iter_rows():
                for cell in row:
                    cell.number_format = u'#,##0.00'

        # Save output to new filename
        wb.save(filename=self.options.output_filepath)

        # Close workbook
        wb.close()

        logging.info('DONE!')


if __name__ == "__main__":
    FORMAT = "%(asctime)-7s | %(levelname)s  | %(message)s"
    logging.basicConfig(format=FORMAT, level=logging.DEBUG, datefmt='%H:%M:%S')
    logging.info('-- Starting --')

    # Set options for VarianzaFormat
    format = VarianzaFormat()
    format.options.font = 'Consolas'
    format.options.header_bg = '0000FF'
    format.options.odd_bg = 'FF0000'
    format.options.even_bg = '00F700'

    # Run the class
    format.run()
